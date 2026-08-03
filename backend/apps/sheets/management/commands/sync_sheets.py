import csv, os, io, re
import requests as req
import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from apps.sheets.models import VideoMetadata

SHEETS_ID = '1Ga5zMekIlVjHKhxkfGBIhAiCh0H3zGYf0TOoMkoctj4'
XLSX_URL  = f'https://docs.google.com/spreadsheets/d/{SHEETS_ID}/export?format=xlsx'

REGISTRO_LINK_COLS = {5, 7, 8, 12, 14}

class Command(BaseCommand):

    def handle(self, *args, **kwargs):
        self.stdout.write("Sincronizando datos (preservando estados de app)...")

        # Contadores: antes los fallos se tragaban con `except Exception: pass`,
        # así que era imposible saber si se habían sincronizado 400 filas o 4.
        self.errores = []

        def _upsert(filtro, defaults, etiqueta):
            """update_or_create tolerante a duplicados, dejando rastro de cada fallo.

            video_id / id_video_equipo NO son únicos en el modelo: si hay duplicados,
            update_or_create lanza MultipleObjectsReturned. Antes eso silenciaba la
            fila entera y el video no aparecía nunca en la web.
            """
            try:
                VideoMetadata.objects.update_or_create(defaults=defaults, **filtro)
                return True
            except VideoMetadata.MultipleObjectsReturned:
                # Hay duplicados en BD. No borramos nada: actualizamos el más antiguo
                # y lo dejamos reportado para poder limpiarlo a mano.
                ids = list(VideoMetadata.objects.filter(**filtro)
                           .order_by('id').values_list('id', flat=True))
                VideoMetadata.objects.filter(pk=ids[0]).update(**defaults)
                self.errores.append(
                    f'{etiqueta}: {len(ids)} filas duplicadas en BD (ids {ids}) — actualizada la primera')
                return True
            except Exception as e:
                self.errores.append(f'{etiqueta}: {type(e).__name__} {e}')
                return False

        def safe_upsert_registro(video_id, defaults):
            """Create or update a registro record without touching app-managed fields."""
            return _upsert({'tipo': 'registro', 'video_id': video_id}, defaults,
                           f'registro {video_id}')

        def safe_upsert_censo(id_equipo, defaults):
            """Create or update a censo record without touching estado_censo/reservado_por."""
            id_equipo = str(id_equipo or '').strip()
            if not id_equipo:
                # Sin ID de equipo, update_or_create hacía match con la MISMA fila
                # vacía una y otra vez: N filas del CSV colapsaban en un solo registro.
                self.errores.append('censo: fila sin ID DE VIDEO EQUIPO — omitida')
                return False
            return _upsert({'tipo': 'censo', 'id_video_equipo': id_equipo}, defaults,
                           f'censo {id_equipo}')

        # ---------------------------------------------------------------
        # REGISTRO — leer desde XLSX para preservar hipervínculos de Smart Chips
        # ---------------------------------------------------------------
        count = 0
        try:
            self.stdout.write("Descargando Registro desde Google Sheets (XLSX)...")
            resp = req.get(XLSX_URL, timeout=60)
            resp.raise_for_status()
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)

            sheet_names = [s.lower() for s in wb.sheetnames]
            if 'registro' in sheet_names:
                ws = wb[wb.sheetnames[sheet_names.index('registro')]]
            else:
                ws = wb.active
            self.stdout.write(f"  Hoja: '{ws.title}' ({ws.max_row} filas)")

            def extraer_celda(cell, col_idx=None):
                val = cell.value
                if isinstance(val, float) and val.is_integer():
                    val = int(val)
                texto = str(val).strip() if val is not None else ''
                if texto.lower() in ('none', ''):
                    texto = ''

                if col_idx in REGISTRO_LINK_COLS:
                    if cell.hyperlink and cell.hyperlink.target:
                        return str(cell.hyperlink.target).strip()
                    m = re.match(r'=HYPERLINK\("([^"]+)"', texto, re.IGNORECASE)
                    if m:
                        return m.group(1)
                    if texto.lower().startswith('http'):
                        return texto
                    return ''

                return texto

            headers = [extraer_celda(c) for c in next(ws.iter_rows(min_row=1, max_row=1))]

            def col_idx(name):
                name_l = name.lower().strip()
                for i, h in enumerate(headers):
                    if h.lower().strip() == name_l:
                        return i
                return None

            idx_id    = col_idx('id')
            idx_mb    = col_idx('miembro')
            idx_mm    = col_idx('mateo/miguel')
            idx_est   = col_idx('estilizado')
            idx_pi    = col_idx('prompt imagen')
            idx_img   = col_idx('imagen')
            idx_pv    = col_idx('prompt video')
            idx_vid   = col_idx('video')
            idx_orig  = col_idx('video orginal') or col_idx('video original')
            idx_acep  = col_idx('aceptado')
            idx_pf    = col_idx('prompt final')

            def cell_val(row_cells, idx, is_link=False):
                if idx is None or idx >= len(row_cells):
                    return ''
                return extraer_celda(row_cells[idx],
                                     col_idx=idx if is_link else None)

            skipped = 0
            for row in ws.iter_rows(min_row=2):
                row_cells = list(row)
                v_id      = cell_val(row_cells, idx_id)
                video_url = cell_val(row_cells, idx_vid, is_link=True)
                img_url   = cell_val(row_cells, idx_img, is_link=True)
                prompt    = cell_val(row_cells, idx_pv)

                if not v_id or not video_url or not img_url or not prompt:
                    skipped += 1
                    continue

                # Only update source fields — never touch estado_revision / comentario_revision
                safe_upsert_registro(v_id, {
                    'usuario':              cell_val(row_cells, idx_mb),
                    'mateo_miguel':         cell_val(row_cells, idx_mm),
                    'estilizado':           cell_val(row_cells, idx_est),
                    'prompt_imagen':        cell_val(row_cells, idx_pi),
                    'imagen_link':          img_url,
                    'prompt_video':         prompt,
                    'drive_link':           video_url,
                    'video_original_link':  cell_val(row_cells, idx_orig, is_link=True),
                    'aceptado':             cell_val(row_cells, idx_acep),
                    'prompt_final':         cell_val(row_cells, idx_pf),
                })
                count += 1

            self.stdout.write(f"Registro sincronizado desde Sheets: {count} filas ({skipped} incompletas omitidas).")

        except Exception as e:
            self.stdout.write(f"XLSX no disponible ({e}), usando registro.csv local...")
            # Ruta absoluta: con la ruta relativa, si el proceso arrancaba desde otro
            # directorio el archivo "no existía" y no se sincronizaba nada.
            registro_path = os.path.join(settings.BASE_DIR, 'registro.csv')
            if os.path.exists(registro_path):
                def get_v(row, *keys):
                    for k in keys:
                        for rk, rv in row.items():
                            if rk and rk.lower().strip() == k.lower().strip():
                                v = rv.strip()
                                if v and not v.startswith('http') and '.' in v and '/' not in v:
                                    return ''
                                return v
                    return ''
                with open(registro_path, mode='r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        v_id = get_v(row, 'Id')
                        if not v_id:
                            continue
                        safe_upsert_registro(v_id, {
                            'usuario':             get_v(row, 'Miembro'),
                            'mateo_miguel':        get_v(row, 'Mateo/Miguel'),
                            'estilizado':          get_v(row, 'Estilizado'),
                            'prompt_imagen':       get_v(row, 'Prompt imagen'),
                            'imagen_link':         get_v(row, 'imagen'),
                            'prompt_video':        get_v(row, 'prompt video'),
                            'drive_link':          get_v(row, 'video'),
                            'video_original_link': get_v(row, 'video orginal', 'video original'),
                            'aceptado':            get_v(row, 'ACEPTADO'),
                            'prompt_final':        get_v(row, 'Prompt Final'),
                        })
                        count += 1
                self.stdout.write(f"Registro sincronizado desde CSV: {count} filas.")
            else:
                self.stdout.write("No se encontro registro.csv")

        # ---------------------------------------------------------------
        # CENSO — CSV local, preserva estado_censo y estados de app
        # ---------------------------------------------------------------
        censo_path = os.path.join(settings.BASE_DIR, 'censo.csv')
        if os.path.exists(censo_path):
            count = 0
            with open(censo_path, mode='r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    v_id     = row.get('ID DE VIDEO', '').strip()
                    id_equip = row.get('ID DE VIDEO EQUIPO', '').strip()
                    if not v_id:
                        continue
                    # Normalize float ids like '1.0' → '1'
                    try:
                        id_equip = str(int(float(id_equip))) if id_equip else id_equip
                    except (ValueError, TypeError):
                        pass

                    # Only update source fields — never touch estado_censo / reservado_por
                    safe_upsert_censo(id_equip, {
                        'video_id':       v_id,
                        'usuario':        row.get('usuario', '').strip(),
                        'drive_link':     row.get('LINK', '').strip(),
                        'mapa':           row.get('MAPA', '').strip(),
                        'genero':         row.get('GENERO', '').strip(),
                        'etnia':          row.get('ETNIA', '').strip(),
                        'duracion':       row.get('DURACION', '').strip(),
                        'camara':         row.get('CAMARA', '').strip(),
                        'especie':        row.get('ESPECIE', '').strip(),
                        'plano':          row.get('PLANO', '').strip(),
                        'interior':       row.get('INTERIOR', '').strip(),
                        'accion':         row.get('ACCION', '').strip(),
                    })
                    count += 1
            self.stdout.write(f"Censo sincronizado: {count} filas.")
        else:
            self.stdout.write(self.style.ERROR(f"No se encontro censo.csv en {censo_path}"))

        # Resumen final: si algo se pierde, tiene que verse en los logs del despliegue.
        if self.errores:
            self.stdout.write(self.style.WARNING(
                f"\n{len(self.errores)} filas con problemas durante la sincronizacion:"))
            for msg in self.errores[:50]:
                self.stdout.write(f"  - {msg}")
            if len(self.errores) > 50:
                self.stdout.write(f"  ... y {len(self.errores) - 50} mas")
        else:
            self.stdout.write(self.style.SUCCESS("\nSincronizacion sin errores."))
